"""Servidor MCP para exponer consultas SQL de solo lectura (SELECT) vía stdio."""

from __future__ import annotations

import difflib
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from urllib.parse import quote_plus

import django
import pandas as pd
import psycopg2
from dotenv import load_dotenv
from mcp.server.fastmcp import FastMCP

BASE_DIR = Path(__file__).resolve().parents[1]
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "sgmsc.settings")
django.setup()

from django.conf import settings
from asgiref.sync import sync_to_async

from asignaciones.models import Asignacion
from salas.models import Sala
from usuarios.models import Usuario

load_dotenv(BASE_DIR / ".env")


def _build_database_url_from_parts() -> str | None:
    """Construye DATABASE_URL a partir de DB_* (como en settings.py).

    Formato: postgresql://user:pass@host:port/dbname
    """
    name = os.getenv("DB_NAME")
    user = os.getenv("DB_USER")
    password = os.getenv("DB_PASSWORD")
    host = os.getenv("DB_HOST")
    port = os.getenv("DB_PORT")

    if not all([name, user, password, host, port]):
        return None

    user_enc = quote_plus(user)
    pass_enc = quote_plus(password)
    return f"postgresql://{user_enc}:{pass_enc}@{host}:{port}/{name}"


DATABASE_URL = os.getenv("DATABASE_URL") or _build_database_url_from_parts()

# 2. Inicializar el servidor FastMCP
# Este nombre es interno, sirve para identificar el servidor en los logs
mcp = FastMCP("Servidor_sgmsc")

# 3. Definir la Herramienta (Tool)
# El decorador @mcp.tool() convierte esta función normal de Python 
# en una herramienta que el LLM (Gemini) puede "ver" y usar.
@mcp.tool()
def consultar_base_datos(sql_query: str) -> str:
    """
    Herramienta exclusiva de SOLO LECTURA para el Administrador del sistema. 
    Ejecuta consultas SQL en la base de datos PostgreSQL de la universidad para extraer información sobre monitores, horas acumuladas, horarios y rendimiento.

    REGLAS ESTRICTAS Y LIMITANTES (CRÍTICO):
    1. OPERACIONES PERMITIDAS: Estás restringido EXCLUSIVAMENTE a operaciones de consulta (`SELECT`).
    2. OPERACIONES PROHIBIDAS: BAJO NINGUNA CIRCUNSTANCIA puedes generar, intentar o ejecutar sentencias que modifiquen los datos o la estructura de la base de datos. Esto incluye, pero no se limita a: `INSERT`, `UPDATE`, `DELETE`, `DROP`, `ALTER`, `TRUNCATE` o ejecución de procedimientos almacenados que alteren estado.
    3. MANEJO DE INTENCIONES INSEGURAS: Si el administrador solicita agregar, borrar o modificar información, NO uses esta herramienta. En su lugar, responde directamente que tu rol y tus permisos están limitados estrictamente a la consulta de información.

    El resultado de esta herramienta será una representación en texto de las filas y columnas obtenidas, úsala para redactar una respuesta natural y precisa al administrador.
    """
    
    if not DATABASE_URL:
        return (
            "Error de configuración: falta DATABASE_URL o las variables DB_NAME/DB_USER/DB_PASSWORD/DB_HOST/DB_PORT."
        )

    # Validaciones de seguridad básicas (para desarrollo)
    query_upper = (sql_query or "").upper()
    if any(forbidden in query_upper for forbidden in ["INSERT", "UPDATE", "DELETE", "DROP", "ALTER", "TRUNCATE"]):
        return "Error de seguridad: Solo se permiten consultas SELECT (solo lectura)."

    conn = None
    try:
        conn = psycopg2.connect(DATABASE_URL)
        with conn.cursor() as cur:
            cur.execute(sql_query)
            columnas = [desc[0] for desc in cur.description] if cur.description else []
            resultados = cur.fetchall()
            filas = [dict(zip(columnas, fila)) for fila in resultados]

        return json.dumps(
            {
                "columns": columnas,
                "rows": filas,
                "row_count": len(filas),
            },
            ensure_ascii=False,
            default=str,
        )

    except psycopg2.Error as e:
        # Si la IA se equivoca en la sintaxis SQL, le devolvemos el error
        # para que "aprenda" y corrija la consulta en un segundo intento.
        return json.dumps({"error": f"Error en la consulta SQL: {e}"}, ensure_ascii=False)
        
    except Exception as e:
        return json.dumps({"error": f"Error inesperado del sistema: {e}"}, ensure_ascii=False)
        
    finally:
        # Siempre asegurar que la conexión se cierre, pase lo que pase
        if conn is not None:
            conn.close()

def _normalizar_data_excel(data_json: str | list | dict) -> list[dict]:
    if isinstance(data_json, list):
        return [row for row in data_json if isinstance(row, dict)]
    if isinstance(data_json, dict):
        if "rows" in data_json and isinstance(data_json["rows"], list):
            return [row for row in data_json["rows"] if isinstance(row, dict)]
        return [data_json]

    parsed = json.loads(data_json)
    if isinstance(parsed, list):
        return [row for row in parsed if isinstance(row, dict)]
    if isinstance(parsed, dict) and isinstance(parsed.get("rows"), list):
        return [row for row in parsed["rows"] if isinstance(row, dict)]
    if isinstance(parsed, dict):
        return [parsed]
    raise ValueError("Los datos para Excel deben ser un JSON de filas o un objeto con `rows`.")


def _sanitizar_nombre_archivo(nombre: str) -> str:
    limpio = re.sub(r"[^A-Za-z0-9_-]+", "_", nombre or "reporte").strip("_")
    return limpio or "reporte"


@mcp.tool()
def generate_excel_report(data_json: str, report_name: str = "reporte"):
    """
    Recibe una lista de diccionarios (ej: resultados de un SELECT)
    Esta herramienta recibe una lista de datos y devuelve un enlace de descarga.
    Úsala cuando el usuario pida un 'informe', 'excel', 'archivo' o cuando los datos sean demasiados para
    leerlos cómodamente en el chat.
    """
    rows = _normalizar_data_excel(data_json)
    df = pd.DataFrame(rows)

    filename = f"{_sanitizar_nombre_archivo(report_name)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    relative_path = os.path.join('reports', filename)
    full_path = os.path.join(settings.MEDIA_ROOT, relative_path)

    os.makedirs(os.path.dirname(full_path), exist_ok=True)

    df.to_excel(full_path, index=False, sheet_name='Datos')

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        return f"Error de configuración: falta instalar openpyxl ({exc})."

    wb = load_workbook(full_path)
    ws = wb["Datos"]
    for i, col in enumerate(df.columns, start=1):
        values = [str(col)] + ["" if pd.isna(v) else str(v) for v in df[col].tolist()]
        max_len = max(len(value) for value in values) + 2
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max_len
    wb.save(full_path)

    relative_url_path = relative_path.replace("\\", "/").replace(os.sep, "/").lstrip("/")
    site_url = (settings.SITE_URL or "http://localhost:8000").rstrip("/")
    media_url = (settings.MEDIA_URL or "/media/").strip("/")
    url_final = f"{site_url}/{media_url}/{relative_url_path}" if media_url else f"{site_url}/{relative_url_path}"
    url_final = re.sub(r"(?<!:)//+", "/", url_final)
    return f"✅ Informe generado con éxito. [📥 Descargar Informe Excel]({url_final})"


@mcp.tool()
async def verificar_conflicto_horario(sala_id: int, dia_semana: str, hora_inicio: str, hora_fin: str, semestre_id: int) -> str:
    """
    Verifica si una sala específica ya tiene una asignación que se cruce con el horario propuesto.
    Úsala SIEMPRE antes de confirmar al usuario que una sala puede ser asignada.
    Retorna un mensaje de confirmación si está libre, o los detalles del conflicto si está ocupada.
    """
    
    dia_semana_val = int(dia_semana) if str(dia_semana).isdigit() else dia_semana

    def _buscar_conflicto():
        return (
            Asignacion.objects.select_related("monitor", "horario", "semestre")
            .filter(
                horario__sala_id=sala_id,
                semestre_id=semestre_id,
                horario__dia_semana=dia_semana_val,
                horario__hora_inicio__lt=hora_fin,
                horario__hora_fin__gt=hora_inicio,
            )
            .order_by("horario__hora_inicio")
            .first()
        )

    try:
        resultado = await sync_to_async(_buscar_conflicto, thread_sensitive=True)()
    except Exception as exc:
        return f"Error al verificar conflicto de horario: {exc}"

    if resultado:
        return (
            "CONFLICTO ENCONTRADO: La sala ya está asignada a "
            f"{resultado.monitor.get_full_name()} en el horario de "
            f"{resultado.horario.hora_inicio} a {resultado.horario.hora_fin}. "
            "No se puede realizar la asignación."
        )
    return "SALA LIBRE: No se encontraron cruces de horario. Es seguro proceder con la asignación."



@mcp.tool()
async def busqueda_difusa_nombres(entidad: str, termino_busqueda: str) -> str:
    """
    Realiza una búsqueda tolerante a errores ortográficos. 
    Úsala cuando una consulta SQL por nombre devuelva vacío.
    'entidad' debe ser "monitor" o "sala". 'termino_busqueda' es la palabra mal escrita.
    """
    
    def _obtener_diccionario():
        if entidad.lower() == "monitor":
            return {
                f"{usuario.first_name} {usuario.last_name}".strip(): usuario.id
                for usuario in Usuario.objects.filter(rol="monitor", is_active=True)
            }
        if entidad.lower() == "sala":
            return {
                f"{sala.nombre} ({sala.codigo})": sala.id_sala
                for sala in Sala.objects.all()
            }
        return None

    try:
        diccionario_bd = await sync_to_async(_obtener_diccionario, thread_sensitive=True)()
    except Exception as exc:
        return f"Error al realizar la busqueda difusa: {exc}"
    if diccionario_bd is None:
        return "Error: Entidad no válida. Debe ser 'monitor' o 'sala'."

    nombres_lista = list(diccionario_bd.keys())

    coincidencias = difflib.get_close_matches(termino_busqueda, nombres_lista, n=3, cutoff=0.5)

    if not coincidencias:
        return f"No se encontró nada ni remotamente parecido a '{termino_busqueda}' en la base de datos."

    # 3. Formatear la respuesta para el LLM
    respuesta = f"Búsqueda difusa para '{termino_busqueda}':\n"
    for nombre_match in coincidencias:
        score = difflib.SequenceMatcher(None, termino_busqueda.lower(), nombre_match.lower()).ratio()
        db_id = diccionario_bd[nombre_match]
        respuesta += f"- Posible coincidencia: '{nombre_match}' (Similitud: {round(score * 100, 2)}%) -> Usa el ID: {db_id}\n"

    return respuesta

# 4. Punto de entrada para ejecutar el servidor
if __name__ == "__main__":
    # Inicia el servidor MCP para escuchar peticiones a través de la terminal (stdio)
    mcp.run()
